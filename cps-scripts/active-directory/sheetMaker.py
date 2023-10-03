# Created by : I. Jared
# Purpose    : Automating creation of student logon sheets (by Michael's request). For ease of copying stuff around, this is part of a 
#              two part process. (Step 1. Python reads a template CSV and makes some Excel workbooks. Step 2. Robocopy sends these to a 
#              fileshare on JH-AZFP called LOGON SHEETS.)
#! Any comments marked with a #! after the character are critical.
import csv
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

"""
TODO:
    * More Robust Logging
    * Exception checks for missing CSV file, or if the name is wrong.
    * Sort the student data by grade level. Michael requested this, but openpyxl
      doesn't have a very good built-in sorting function. I'll have to mess with
      pandas and see if I can sort the CSVs before importing them into Excel sheets.
"""

def main():

    # Converting Powerschool school numbers to abbreviations / appropriate mappings. 
    schoolCodeDict = {
        "705" : "HS",
        "710" : "IH",
        "715" : "CA", # the Academy building
        "610" : "JH",
        "210" : "HI",
        "205" : "MI",
        "105" : "CE",
        "110" : "NW",
        "115" : "SS",
    }

    # Adding this extra dict because I'm lazy and don't want to have to put in names in
    # more than one place.
    schoolTitleDict = {
        "HS" : "High School",
        "IH" : "Intermediate High School",
        "CA" : "Academy", 
        "JH" : "Junior High",
        "HI" : "Heritage",
        "MI" : "Mission",
        "CE" : "Central",
        "NW" : "Northwest",
        "SS" : "Southside",
    }

    #! Change this to wherever the script is executing from, ideally in the same location as the
    #! initial data export. THIS PATH IS FOR THE PS EXPORTS AND INITIAL TEMPLATES ONLY
    directoryPath = "C:\\logonSheetScript\\"
    templateCSVpath = "C:\\destinysftp\\studentLogonTemplate.csv"

    # Create the logon sheet templates for each site, overwriting if they exist
    for abbreviations in schoolCodeDict.values():
        filePath = os.path.join(directoryPath, f"{abbreviations} Logon Sheet.csv")

        # Remove the file if it exists
        if os.path.exists(filePath):
            os.remove(filePath)

        with open(filePath, 'w') as file:
            file.write("First_Name,Last_Name,Student_Number,Grade_Level,School_ID,Email_Address,Google_Password,Microsoft_Password\n")

    #! Based off of the template file name in PS. Do not change without doing it in PS first
    with open(templateCSVpath, mode='r') as file:
        csvFile = csv.reader(file)
    
        for row in csvFile:
            #! DO NOT MESS WITH THESE COLUMN HEADERS. THEY ARE BASED OFF OF THE studentLogonTemplate TEMPLATE IN PS. CHANGE THE 
            #! TEMPLATE BEFORE YOU CHANGE ANYTHING HERE!
            studentNumber = row[0].strip()
            firstName = row[1].strip()
            lastName = row[2].strip()
            emailAddress = row[3].strip() # Leaving EXACTLY how it is in Powerschool, that way we know what accounts to fix.
            gradeLevel = row[4].strip()
            # Had to use a placeholder since Python was complaining otherwise.
            # ERROR is the default placeholder for bad data.
            schoolIDholder = row[5].strip()
            schoolID = schoolCodeDict.get(schoolIDholder, "ERROR") 
			
			# Needed to ensure the password matches what's in Clever.
			# Also Michael didn't want the apostrophes in first / last names on the sheets.
            firstName = firstName.replace("'", "")
            lastName = lastName.replace("'", "")
            
            # Running data checks for each student before processing them into a logon sheet.
            if firstName and lastName and studentNumber:   

                # Have to add this error handle since Michael is completely terrified of feeding email CSVs back into Powerschool for whatever reason.
                if len(emailAddress) == 0:
                    print(f"ERROR: EMAIL ADDRESS FIELD IS EMPTY FOR ({firstName} {lastName}) Adding to ERROR Logon file!")
                    # TODO: Actually add this feature to an error file.

                else: 
                    # Making these two separate vars in case we decide to change the password
                    # convention in the future.
                    googleUser = f"{firstName[0]}{lastName[0]}{lastName[1]}{studentNumber}"
                    windowsUser = f"{firstName[0]}{lastName[0]}{lastName[1]}{studentNumber}!"

                    studentStringForSite = f"{firstName},{lastName},{studentNumber},{gradeLevel},{schoolID},{emailAddress},{googleUser},{windowsUser}" 
                    print(studentStringForSite)

                    # Handling empty school codes as well as any that aren't our standard 9 sites.
                    if schoolID == "ERROR":
                        print(f"ERROR: INVALID SITE CODE FOR ({emailAddress}) Skipping the processing of this user!")

                    else:
                        # Open the file based on the schoolID and write to it

                        file_name = os.path.join(directoryPath, f"{schoolID} Logon Sheet.csv")
                        with open(file_name, 'a') as templateFile: 
                            templateFile.write(f"{studentStringForSite}\n")
                
            else:
                print("ERROR: MISSING CRUCIAL DATA FOR STUDENT, skipping processing this row!")
    
    """
    *************************************
    *   SECTION FOR CONVERTING THE CSV  * 
    *   TEMPLATES TO EXCEL WORKBOOKS    *
    *************************************
    """
    for abbreviations in schoolCodeDict.values():
        csvTemplatePath = os.path.join(directoryPath, f"{abbreviations} Logon Sheet.csv")
        excelFilePath = os.path.join(directoryPath, f"{abbreviations} Logon Sheet.xlsx")
        
        # Check if the CSV file exists
        if os.path.exists(csvTemplatePath):
            # Create a new Excel workbook
            workbook = Workbook()
            worksheet = workbook.active

            # Read the data from the CSV and write it to the Excel worksheet
            with open(csvTemplatePath, 'r') as csvTemplate:
                csvReader = csv.reader(csvTemplate)
                for row in csvReader:
                    worksheet.append(row)

            schoolTitle = schoolTitleDict.get(abbreviations, "ERROR")
            worksheet.insert_rows(1)
            worksheet.merge_cells('A1:H1')
            titleCell = worksheet.cell(row=1, column=1)
            titleCell.font = Font(name="Arial", size=22, bold=True)
            titleCell.value = f"{schoolTitle} Student Logons"
            #orange_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
            titleCell.fill = PatternFill(start_color='ff5800', end_color='ff5800', fill_type='solid')

            # I'm sure there's a less verbose way to do this, but my main thing is getting this done
            # before Michael's deadline. Streamline it later. This formats the cells to be more
			# readable for end-users.
            firstNameCell = worksheet.cell(row=2, column=1)
            firstNameCell.font = Font(name="Arial", size=11, bold=True)
            firstNameCell.value = "First Name"
            firstNameCell.fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
            
            lastNameCell = worksheet.cell(row=2, column=2)
            lastNameCell.font = Font(name="Arial", size=11, bold=True)
            lastNameCell.value = "Last Name" 
            lastNameCell.fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')

            studentNumCell = worksheet.cell(row=2, column=3)
            studentNumCell.font = Font(name="Arial", size=11, bold=True)
            studentNumCell.value = "Student Number"
            studentNumCell.fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid') 

            gradeCell = worksheet.cell(row=2, column=4)
            gradeCell.font = Font(name="Arial", size=11, bold=True)
            gradeCell.value = "Grade Level"
            gradeCell.fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid') 

            siteCell = worksheet.cell(row=2, column=5)
            siteCell.font = Font(name="Arial", size=11, bold=True)
            siteCell.value = "School Code" 
            siteCell.fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')

            emailCell = worksheet.cell(row=2, column=6)
            emailCell.font = Font(name="Arial", size=11, bold=True)
            emailCell.value = "Email Address" 
            emailCell.fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')

            googleCell = worksheet.cell(row=2, column=7)
            googleCell.font = Font(name="Arial", size=11, bold=True)
            googleCell.value = "Google Password" 
            googleCell.fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')

            microsoftCell = worksheet.cell(row=2, column=8)
            microsoftCell.font = Font(name="Arial", size=11, bold=True)
            microsoftCell.value = "Microsoft Password" 
            microsoftCell.fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')

            # Since the column content is usually too long for the default lengths, 
            # this is added here to make it easier for the user to read along with
            # standardizing the appearance of the sheets.
            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)  # Get the column letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 4)  # Add padding
                worksheet.column_dimensions[column_letter].width = adjusted_width

            workbook.save(excelFilePath)

            print(f"Converted {csvTemplatePath} to {excelFilePath}")

    # Adding this here to ensure we have a consistent filename for the PS export; I don't want to have to add
    # extra logic for separate dates. As for the other files, I'll just overwrite them each time so deleting
    # them is unnecessary.
    os.remove(templateCSVpath)

if __name__ == "__main__":
    main()
